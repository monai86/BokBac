import os
import json
import openpyxl
from openpyxl.styles import Font, Alignment

def write_predictions():
    json_path = 'validation/temp_predictions.json'
    excel_path = 'validation/BokBac_validation_template_with_reference_profiles.xlsx'

    if not os.path.exists(json_path):
        print(f"Error: {json_path} does not exist!")
        return

    # Load predictions JSON
    with open(json_path, 'r', encoding='utf-8') as f:
        predictions = json.load(f)

    # Convert predictions list to a lookup dict by expected_organism name
    pred_map = {pred['expected_organism']: pred for pred in predictions}

    # Load Excel Workbook
    wb = openpyxl.load_workbook(excel_path)
    ws = wb['Reference_Profile_Input']

    # Get header row to map columns
    headers = [cell.value for cell in ws[1]]
    col_mapping = {header: idx for idx, header in enumerate(headers, 1)}

    # Verify column existence
    required_cols = [
        'Expected Organism (ID)', 'bokbac_top1', 'bokbac_top2', 
        'bokbac_top3', 'bokbac_confidence', 'top1_correct', 
        'top3_correct', 'review_note'
    ]
    for col in required_cols:
        if col not in col_mapping:
            print(f"Error: Column '{col}' not found in sheet headers!")
            return

    expected_org_idx = col_mapping['Expected Organism (ID)']
    top1_idx = col_mapping['bokbac_top1']
    top2_idx = col_mapping['bokbac_top2']
    top3_idx = col_mapping['bokbac_top3']
    conf_idx = col_mapping['bokbac_confidence']
    top1_corr_idx = col_mapping['top1_correct']
    top3_corr_idx = col_mapping['top3_correct']
    rev_note_idx = col_mapping['review_note']

    body_font = Font(name='Calibri', size=11)
    center_align = Alignment(horizontal='center')
    left_align = Alignment(horizontal='left')

    top1_correct_count = 0
    top3_correct_count = 0
    total_cases = 0

    # Scan rows starting from row 2
    for r in range(2, 500):
        expected_org_cell = ws.cell(row=r, column=expected_org_idx).value
        if not expected_org_cell:
            # End of table
            break

        total_cases += 1
        pred = pred_map.get(expected_org_cell)

        if pred:
            ws.cell(row=r, column=top1_idx, value=pred['bokbac_top1']).font = body_font
            ws.cell(row=r, column=top1_idx).alignment = left_align
            
            ws.cell(row=r, column=top2_idx, value=pred['bokbac_top2']).font = body_font
            ws.cell(row=r, column=top2_idx).alignment = left_align

            ws.cell(row=r, column=top3_idx, value=pred['bokbac_top3']).font = body_font
            ws.cell(row=r, column=top3_idx).alignment = left_align

            ws.cell(row=r, column=conf_idx, value=pred['bokbac_confidence']).font = body_font
            ws.cell(row=r, column=conf_idx).alignment = center_align

            ws.cell(row=r, column=top1_corr_idx, value=pred['top1_correct']).font = body_font
            ws.cell(row=r, column=top1_corr_idx).alignment = center_align

            ws.cell(row=r, column=top3_corr_idx, value=pred['top3_correct']).font = body_font
            ws.cell(row=r, column=top3_corr_idx).alignment = center_align

            ws.cell(row=r, column=rev_note_idx, value=pred['review_note']).font = body_font
            ws.cell(row=r, column=rev_note_idx).alignment = left_align

            if pred['top1_correct'] == 'Yes':
                top1_correct_count += 1
            if pred['top3_correct'] == 'Yes':
                top3_correct_count += 1
        else:
            print(f"Warning: No prediction found for expected organism '{expected_org_cell}' in row {r}!")

    # Save Excel
    wb.save(excel_path)
    print(f"Populated validation results successfully saved to {excel_path}")

    # Remove temporary JSON file
    os.remove(json_path)

    # Print accuracy reports
    top1_accuracy = (top1_correct_count / total_cases) * 100 if total_cases > 0 else 0
    top3_accuracy = (top3_correct_count / total_cases) * 100 if total_cases > 0 else 0

    print("\n" + "="*50)
    print("      BOKBAC ACADEMIC VALIDATION RESULTS REPORT      ")
    print("="*50)
    print(f"Total reference cases validated: {total_cases}")
    print(f"Top-1 Correct predictions:       {top1_correct_count} ({round(top1_accuracy, 2)}%)")
    print(f"Top-3 Correct predictions:       {top3_correct_count} ({round(top3_accuracy, 2)}%)")
    print("="*50)

if __name__ == '__main__':
    write_predictions()
