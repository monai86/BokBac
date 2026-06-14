import os
import json
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def populate_profiles():
    temp_dir = 'validation/temp'
    template_path = 'validation/BokBac_validation_template.xlsx'
    output_path = 'validation/BokBac_validation_template_with_reference_profiles.xlsx'
    csv_path = 'validation/reference_profiles_generated.csv'

    # Load JSON data
    with open(os.path.join(temp_dir, 'tests.json'), 'r', encoding='utf-8') as f:
        tests = json.load(f)
    with open(os.path.join(temp_dir, 'library.json'), 'r', encoding='utf-8') as f:
        library = json.load(f)

    # 1. Load the existing template
    wb = openpyxl.load_workbook(template_path)

    # 2. Update Allowed_Values sheet to use Organism Names instead of Organism IDs
    ws_allowed = wb['Allowed_Values']
    organism_names = sorted([bug['name'] for bug in library])
    
    # Clear Column F (Organism_IDs / Names) starting from row 2
    for r in range(2, 500):
        ws_allowed.cell(row=r, column=6, value=None)
    
    # Write names
    for row_idx, name in enumerate(organism_names, 2):
        ws_allowed.cell(row=row_idx, column=6, value=name).font = Font(name='Calibri', size=11)

    # 3. Populate Reference_Profile_Input sheet
    ws_ref = wb['Reference_Profile_Input']

    # Get header row to map columns correctly
    headers = [cell.value for cell in ws_ref[1]]
    col_mapping = {header: idx for idx, header in enumerate(headers, 1)}

    # Key to Header label map
    key_to_header = {
        'case_id': 'Case ID',
        'dataset_type': 'Dataset Type',
        'gram_reaction': 'Gram Reaction',
        'morphology': 'Morphology',
        'arrangement': 'Arrangement',
        'expected_organism': 'Expected Organism (ID)',
        'reference_source': 'Reference Source',
        'included_in_analysis': 'Included In Analysis',
        'exclusion_reason': 'Exclusion Reason',
        'note': 'Note'
    }

    # Map biochem labels/names to test IDs
    label_to_id = {}
    for t in tests:
        label_to_id[t['label'].lower().replace(' ', '')] = t['id']
        label_to_id[t['id'].lower().replace('_', '')] = t['id']

    def find_test_id(name):
        n = name.lower().replace(' ', '').replace('_', '').replace('₂', '2').replace('₃', '3').replace('°', '')
        
        # 1. Direct mapping
        if n in label_to_id:
            return label_to_id[n]
            
        # 2. Intelligent pattern matching
        if 'ld' in n and '1' in n and 'nacl' in n:
            return 'ldc'
        if 'od' in n and '1' in n and 'nacl' in n:
            return 'odc'
        if 'ad' in n and '1' in n and 'nacl' in n:
            return 'adh'
        if 'xfactor' in n or 'factorx' in n or 'hemin' in n:
            return 'factor_x'
        if 'vfactor' in n or 'factorv' in n or 'nad' in n:
            return 'factor_v'
        if 'satellit' in n:
            return 'satellitism'
        if 'growth' in n and '42' in n:
            return 'growth_42'
        if 'growth' in n and '45' in n:
            return 'growth_45'
        if 'mannitol' in n or 'msa' in n:
            return 'mannitol'
        if 'ornithine' in n or 'odc' in n or 'od(' in n:
            return 'odc'
        if 'lysine' in n or 'ldc' in n:
            return 'ldc'
        if 'arginine' in n or 'adh' in n:
            return 'adh'
        if 'pyr' in n:
            return 'pyr'
        if 'ppr' in n:
            return 'ppr'
        if 'esculin' in n:
            return 'esculin'
        if 'solubility' in n:
            return 'bile_solubility'
        if 'gas' in n and 'glucose' in n:
            return 'gas_glucose'
        if 'macconkey' in n or 'mac' in n:
            return 'growth_macconkey'
        if 'xld' in n:
            return 'colony_xld'
        if 'tcbs' in n:
            return 'tcbs'
        if 'h2s' in n:
            return 'h2s'
        if 'gas' in n:
            return 'gas_glucose'
        if 'tsi' in n:
            return 'tsi'
        if 'nacl6' in n or 'salt6' in n or '6.5%nacl' in n or '6.5nacl' in n:
            return 'salt_6'
        if 'nacl8' in n or 'salt8' in n:
            return 'salt_8'
        if 'nacl0' in n or 'salt0' in n:
            return 'salt_0'
        if 'nacl1' in n or 'salt1' in n:
            return 'salt_1'
        if 'nacl10' in n or 'salt10' in n:
            return 'salt_10'
        if 'optochin' in n:
            return 'optochin'
        if 'bacitracin' in n:
            return 'bacitracin'
        if 'novobiocin' in n:
            return 'novobiocin'
        if 'coagulase' in n:
            return 'coagulase'
        if 'catalase' in n:
            return 'catalase'
        if 'oxidase' in n:
            return 'oxidase'
        if 'dnase' in n:
            return 'dnase'
        if 'hemolysis' in n:
            return 'hemolysis'
        if 'camp' in n:
            return 'camp'
        if 'urease' in n or 'urea' in n:
            return 'urease'
        if 'motility' in n or 'motile' in n:
            return 'motility'
        if 'citrate' in n:
            return 'citrate'
        if 'indole' in n:
            return 'indole'
        if 'mr' in n:
            return 'mr'
        if 'vp' in n:
            return 'vp'
        if 'gelatin' in n:
            return 'gelatin'
        if 'starch' in n:
            return 'starch'
        if 'onpg' in n:
            return 'onpg'
        if 'nitrate' in n:
            return 'nitrate'
        if 'malonate' in n:
            return 'malonate'
        if 'kcn' in n:
            return 'kcn'
        if 'acetamide' in n:
            return 'acetamide'
        if 'cetrimide' in n:
            return 'cetrimide'
        if 'o129' in n:
            return 'o129'
        if 'string' in n:
            return 'string_test'
        if 'glucose' in n:
            return 'glucose'
        if 'lactose' in n:
            return 'lactose'
        if 'sucrose' in n:
            return 'sucrose'
        if 'maltose' in n:
            return 'maltose'
        if 'trehalose' in n:
            return 'trehalose'
        if 'arabinose' in n:
            return 'arabinose'
        if 'sorbitol' in n:
            return 'sorbitol'
        if 'xylose' in n:
            return 'xylose'
        if 'fructose' in n:
            return 'fructose'
        if 'mannose' in n:
            return 'mannose'
        if 'adonitol' in n:
            return 'adonitol'
        if 'dulcitol' in n:
            return 'dulcitol'
        if 'inositol' in n:
            return 'inositol'
        if 'raffinose' in n:
            return 'raffinose'
        if 'rhamnose' in n:
            return 'rhamnose'
        if 'salicin' in n:
            return 'salicin'
        if 'kingf' in n:
            return 'king_f'
        if 'kingp' in n:
            return 'king_p'
        if 'lecithinase' in n:
            return 'lecithinase'
        return None

    # Helper: Map Gram and Morph from database
    def map_gram(gram):
        g = String_Clean(gram)
        if g == '+': return 'positive'
        if g == '-': return 'negative'
        return 'variable'

    def map_morph(morph):
        m = String_Clean(morph).lower()
        if 'coccobacill' in m: return 'coccobacilli'
        if 'cocci' in m or 'coccus' in m: return 'cocci'
        if 'curved' in m or 'comma' in m or 'vibrio' in m: return 'curved_rod'
        if 'filament' in m or 'branching' in m: return 'branching_filament'
        if 'rod' in m or 'bacill' in m: return 'bacilli'
        return 'unknown'

    def map_arrangement(morph):
        m = String_Clean(morph).lower()
        if 'cluster' in m: return 'cluster'
        if 'chain' in m: return 'chain'
        if 'pair' in m: return 'pairs'
        if 'diplococci' in m or 'diplo' in m: return 'diplococci'
        if 'palisade' in m: return 'palisade'
        if 'single' in m: return 'single'
        return 'unknown'

    def String_Clean(val):
        return str(val or '').replace('−', '-').replace('—', '-').strip()

    body_font = Font(name='Calibri', size=11)
    center_align = Alignment(horizontal='center')
    left_align = Alignment(horizontal='left')

    csv_rows = []
    
    # Pre-fill all 1000 rows with 'internal_reference_profile' in the dataset_type column
    dataset_type_col_idx = col_mapping.get('Dataset Type')
    if dataset_type_col_idx:
        for r in range(2, 1001):
            ws_ref.cell(row=r, column=dataset_type_col_idx, value='internal_reference_profile').font = body_font

    # Keep track of statistics
    organisms_included = []
    missing_test_data = {} # test_id -> count of missing values
    total_populated_columns = set()
    unmapped_profiles = []

    for idx, bug in enumerate(library, 1):
        case_id = f"REF{idx:03d}"
        org_name = bug['name']
        organisms_included.append(org_name)

        gram = bug.get('gram', '')
        morph = bug.get('morph', '')

        gram_mapped = map_gram(gram)
        morph_mapped = map_morph(morph)
        arr_mapped = map_arrangement(morph)

        row_data = {
            'case_id': case_id,
            'dataset_type': 'internal_reference_profile',
            'gram_reaction': gram_mapped,
            'morphology': morph_mapped,
            'arrangement': arr_mapped,
            'expected_organism': org_name,
            'reference_source': 'BokBac internal organism database',
            'included_in_analysis': 'Yes',
            'exclusion_reason': '',
            'note': bug.get('notes', '')
        }

        # Populate biochemical profile tests
        bug_biochem = bug.get('biochem', [])
        profile_map = {}
        for b in bug_biochem:
            tid = find_test_id(b['t'])
            if tid:
                profile_map[tid] = b['r']
                total_populated_columns.add(f"test__{tid}")
            else:
                unmapped_profiles.append((org_name, b['t'], b['r']))

        # Merge profile map into row data
        for test in tests:
            col_name = f"test__{test['id']}"
            val = profile_map.get(test['id'], '')
            row_data[col_name] = val
            if not val:
                missing_test_data[test['id']] = missing_test_data.get(test['id'], 0) + 1

        # Write row to sheet using header labels instead of raw keys
        current_row = idx + 1 # offset header
        for col_key, val in row_data.items():
            # Resolve actual header label
            header_label = key_to_header.get(col_key, col_key)
            col_idx = col_mapping.get(header_label)
            if col_idx:
                cell = ws_ref.cell(row=current_row, column=col_idx, value=val)
                cell.font = body_font
                if col_key.startswith('test__') or col_key in ['gram_reaction', 'morphology', 'arrangement', 'included_in_analysis']:
                    cell.alignment = center_align
                else:
                    cell.alignment = left_align

        # Build CSV row
        csv_row = [row_data.get(headers_dict_key, '') for headers_dict_key in [
            'case_id', 'dataset_type', 'gram_reaction', 'morphology', 'arrangement',
            'expected_organism', 'reference_source', 'included_in_analysis', 'exclusion_reason', 'note'
        ]] + [row_data.get(f"test__{t['id']}", '') for t in tests] + [''] * 7 # blank output columns
        csv_rows.append(csv_row)

    # Save Excel
    wb.save(output_path)
    print(f"Populated workbook successfully saved to {output_path}")

    # Write CSV
    with open(csv_path, 'w', encoding='utf-8', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        writer.writerows(csv_rows)
    print(f"CSV exported successfully to {csv_path}")

    # Print summary report
    print("\n--- POPULATION SUMMARY REPORT ---")
    print(f"1. Number of reference profiles generated: {len(library)}")
    print(f"2. Number of biochem test columns populated: {len(total_populated_columns)} (out of {len(tests)} total tests)")
    print(f"3. Organisms count: {len(organisms_included)}")
    
    # Identify tests that had missing data
    print("4. Tests with missing data:")
    sorted_missing = sorted(missing_test_data.items(), key=lambda x: x[1], reverse=True)
    for tid, count in sorted_missing[:10]:
        print(f"   - {tid}: {count} profiles missing ({round(count/len(library)*100, 1)}%)")
    if len(sorted_missing) > 10:
        print(f"   - ... and {len(sorted_missing) - 10} other tests.")
        
    print("5. Unmapped profile entries:")
    if unmapped_profiles:
        for org, label, r_val in unmapped_profiles:
            print(f"   - {org}: '{label}' -> '{r_val}' (No matching test ID found)")
    else:
        print("   - None! All biochemical profile tests mapped cleanly.")

if __name__ == '__main__':
    populate_profiles()
