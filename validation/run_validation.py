import os
import json
import csv
import subprocess
import openpyxl
from openpyxl.styles import Font, Alignment

# Paths
EXCEL_PATH = 'validation/BokBac_validation_template_with_reference_profiles.xlsx'
TEMP_INPUT_JSON = 'validation/temp/cases_to_predict.json'
TEMP_OUTPUT_JSON = 'validation/temp/predictions_output.json'
LIBRARY_JSON_PATH = 'validation/temp/library.json'

def is_missing_value(val):
    if val is None:
        return True
    s = str(val).strip()
    if s == "" or s.lower() in ["not done", "unknown", "—", "–", "-", "n/a"]:
        return True
    return False

def parse_excel_sheet(ws):
    headers = [cell.value for cell in ws[1]]
    col_mapping = {header: idx for idx, header in enumerate(headers, 1)}
    
    # Required columns for metadata
    case_id_col = 'Case ID'
    dataset_type_col = 'Dataset Type'
    gram_col = 'Gram Reaction'
    morph_col = 'Morphology'
    arr_col = 'Arrangement'
    inc_col = 'Included In Analysis'
    
    # Determine expected organism column
    if 'Expected Organism (ID)' in col_mapping:
        expected_org_col = 'Expected Organism (ID)'
    elif 'Final Identification (ID)' in col_mapping:
        expected_org_col = 'Final Identification (ID)'
    else:
        raise ValueError(f"No Expected Organism or Final Identification column found in sheet {ws.title}")
        
    specimen_col = 'Specimen Type' if 'Specimen Type' in col_mapping else None
    
    cases = []
    
    # Read rows starting from row 2
    for r in range(2, 2000):
        # Stop if row is completely empty or Case ID is empty
        case_id = ws.cell(row=r, column=col_mapping[case_id_col]).value
        if not case_id:
            break
            
        included = ws.cell(row=r, column=col_mapping[inc_col]).value
        if included != 'Yes':
            continue
            
        expected_org = ws.cell(row=r, column=col_mapping[expected_org_col]).value
        gram = ws.cell(row=r, column=col_mapping[gram_col]).value
        morph = ws.cell(row=r, column=col_mapping[morph_col]).value
        arr = ws.cell(row=r, column=col_mapping[arr_col]).value
        dataset_type = ws.cell(row=r, column=col_mapping[dataset_type_col]).value
        specimen = ws.cell(row=r, column=col_mapping[specimen_col]).value if specimen_col else None
        
        # Extract biochemical tests (test__*)
        answers = {}
        for h, idx in col_mapping.items():
            if h and h.startswith('test__'):
                test_id = h.replace('test__', '')
                cell_val = ws.cell(row=r, column=idx).value
                if not is_missing_value(cell_val):
                    answers[test_id] = str(cell_val).strip()
                    
        cases.append({
            'row_idx': r,
            'case_id': str(case_id).strip(),
            'dataset_type': str(dataset_type).strip() if dataset_type else "",
            'expected_organism': str(expected_org).strip() if expected_org else "",
            'gram_reaction': str(gram).strip() if gram else "",
            'morphology': str(morph).strip() if morph else "",
            'arrangement': str(arr).strip() if arr else "",
            'specimen_type': str(specimen).strip() if specimen else "",
            'answers': answers
        })
        
    return cases, col_mapping

def write_predictions_to_sheet(ws, predictions, col_mapping):
    body_font = Font(name='Calibri', size=11)
    center_align = Alignment(horizontal='center')
    left_align = Alignment(horizontal='left')
    
    # Map predictions by case_id
    pred_by_id = {p['case_id']: p for p in predictions}
    
    # Get column indices
    top1_idx = col_mapping['bokbac_top1']
    top2_idx = col_mapping['bokbac_top2']
    top3_idx = col_mapping['bokbac_top3']
    conf_idx = col_mapping['bokbac_confidence']
    top1_corr_idx = col_mapping['top1_correct']
    top3_corr_idx = col_mapping['top3_correct']
    rev_note_idx = col_mapping['review_note']
    
    for r in range(2, 2000):
        case_id = ws.cell(row=r, column=col_mapping['Case ID']).value
        if not case_id:
            break
            
        case_id_str = str(case_id).strip()
        if case_id_str in pred_by_id:
            pred = pred_by_id[case_id_str]
            
            ws.cell(row=r, column=top1_idx, value=pred['predicted_top1']).font = body_font
            ws.cell(row=r, column=top1_idx).alignment = left_align
            
            ws.cell(row=r, column=top2_idx, value=pred['predicted_top2']).font = body_font
            ws.cell(row=r, column=top2_idx).alignment = left_align
            
            ws.cell(row=r, column=top3_idx, value=pred['predicted_top3']).font = body_font
            ws.cell(row=r, column=top3_idx).alignment = left_align
            
            ws.cell(row=r, column=conf_idx, value=pred['confidence_score']).font = body_font
            ws.cell(row=r, column=conf_idx).alignment = center_align
            
            ws.cell(row=r, column=top1_corr_idx, value=pred['top1_correct']).font = body_font
            ws.cell(row=r, column=top1_corr_idx).alignment = center_align
            
            ws.cell(row=r, column=top3_corr_idx, value=pred['top3_correct']).font = body_font
            ws.cell(row=r, column=top3_corr_idx).alignment = center_align
            
            ws.cell(row=r, column=rev_note_idx, value=pred['warning_note']).font = body_font
            ws.cell(row=r, column=rev_note_idx).alignment = left_align

def compute_metrics(predictions):
    total = len(predictions)
    if total == 0:
        return {
            'included_cases': 0,
            'top1_accuracy': 0.0,
            'top3_accuracy': 0.0,
            'precision': 0.0,
            'recall': 0.0,
            'f1_score': 0.0,
            'macro_f1': 0.0
        }
        
    top1_correct = sum(1 for p in predictions if p['top1_correct'] == 'Yes')
    top3_correct = sum(1 for p in predictions if p['top3_correct'] == 'Yes')
    
    top1_acc = (top1_correct / total) * 100
    top3_acc = (top3_correct / total) * 100
    
    # Calculate precision, recall, f1 per organism class
    classes = set(p['expected_organism'] for p in predictions)
    
    precision_list = []
    recall_list = []
    f1_list = []
    
    for c in classes:
        tp = sum(1 for p in predictions if p['expected_organism'] == c and p['predicted_top1'] == c)
        fp = sum(1 for p in predictions if p['expected_organism'] != c and p['predicted_top1'] == c)
        fn = sum(1 for p in predictions if p['expected_organism'] == c and p['predicted_top1'] != c)
        
        prec = tp / (tp + fp) if (tp + fp) > 0 else 0.0
        rec = tp / (tp + fn) if (tp + fn) > 0 else 0.0
        f1 = (2 * prec * rec) / (prec + rec) if (prec + rec) > 0 else 0.0
        
        precision_list.append(prec)
        recall_list.append(rec)
        f1_list.append(f1)
        
    macro_prec = (sum(precision_list) / len(classes)) * 100 if classes else 0.0
    macro_rec = (sum(recall_list) / len(classes)) * 100 if classes else 0.0
    macro_f1 = (sum(f1_list) / len(classes)) * 100 if classes else 0.0
    
    return {
        'included_cases': total,
        'top1_accuracy': top1_acc,
        'top3_accuracy': top3_acc,
        'precision': macro_prec,
        'recall': macro_rec,
        'f1_score': macro_f1,  # User asked for "F1-score" and "Macro-F1 score". We provide both.
        'macro_f1': macro_f1
    }

def generate_organism_confusion_matrix(predictions, output_csv_path):
    # Collect all unique expected organisms and predicted top1 organisms
    expected_orgs = sorted(list(set(p['expected_organism'] for p in predictions)))
    predicted_orgs = sorted(list(set(p['predicted_top1'] for p in predictions)))
    
    # Columns of confusion matrix should include all unique expected + predicted
    all_cols = sorted(list(set(expected_orgs + predicted_orgs)))
    
    # Initialize matrix
    matrix = {expected: {pred: 0 for pred in all_cols} for expected in expected_orgs}
    
    # Populate matrix
    for p in predictions:
        matrix[p['expected_organism']][p['predicted_top1']] += 1
        
    # Write to CSV
    with open(output_csv_path, 'w', encoding='utf-8', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(['True \\ Predicted'] + all_cols)
        for expected in expected_orgs:
            row = [expected] + [matrix[expected][pred] for pred in all_cols]
            writer.writerow(row)

def generate_group_confusion_matrix(predictions, library_lookup, output_csv_path):
    # Map predictions to groups
    # Look up groups
    unique_groups = set()
    mapped_predictions = []
    
    for p in predictions:
        true_name = p['expected_organism'].lower()
        pred_name = p['predicted_top1'].lower()
        
        true_group = library_lookup.get(true_name, 'unknown')
        pred_group = library_lookup.get(pred_name, 'unknown')
        
        unique_groups.add(true_group)
        unique_groups.add(pred_group)
        
        mapped_predictions.append((true_group, pred_group))
        
    all_groups = sorted(list(unique_groups))
    
    # Initialize matrix
    matrix = {true_g: {pred_g: 0 for pred_g in all_groups} for true_g in all_groups}
    
    # Populate matrix
    for true_g, pred_g in mapped_predictions:
        matrix[true_g][pred_g] += 1
        
    # Write to CSV
    with open(output_csv_path, 'w', encoding='utf-8', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(['True Group \\ Predicted Group'] + all_groups)
        for true_g in all_groups:
            row = [true_g] + [matrix[true_g][pred_g] for pred_g in all_groups]
            writer.writerow(row)

def export_results_csv(predictions, output_csv_path):
    headers = [
        'case_id', 'dataset_type', 'expected_organism', 'predicted_top1',
        'predicted_top2', 'predicted_top3', 'confidence_score',
        'top1_correct', 'top3_correct', 'warning_note'
    ]
    with open(output_csv_path, 'w', encoding='utf-8', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        for p in predictions:
            writer.writerow([p.get(h, '') for h in headers])

def export_metrics_csv(metrics, total_rows, output_csv_path):
    headers = ['Metric', 'Value']
    with open(output_csv_path, 'w', encoding='utf-8', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        writer.writerow(['Total Rows in Sheet', total_rows])
        writer.writerow(['Included Cases', metrics['included_cases']])
        writer.writerow(['Excluded Cases', total_rows - metrics['included_cases']])
        writer.writerow(['Top-1 Accuracy (%)', round(metrics['top1_accuracy'], 2)])
        writer.writerow(['Top-3 Accuracy (%)', round(metrics['top3_accuracy'], 2)])
        writer.writerow(['Macro-Precision (%)', round(metrics['precision'], 2)])
        writer.writerow(['Macro-Recall (%)', round(metrics['recall'], 2)])
        writer.writerow(['Macro-F1-score (%)', round(metrics['macro_f1'], 2)])

def main():
    print("Step 1: Reading Excel workbook...")
    wb = openpyxl.load_workbook(EXCEL_PATH)
    
    ref_sheet = wb['Reference_Profile_Input']
    real_sheet = wb['Real_Case_Input']
    
    ref_cases, ref_mapping = parse_excel_sheet(ref_sheet)
    real_cases, real_mapping = parse_excel_sheet(real_sheet)
    
    print(f"Loaded {len(ref_cases)} reference cases and {len(real_cases)} real cases for prediction.")
    
    # Save inputs for Vitest runner
    input_payload = {
        'reference_cases': ref_cases,
        'real_cases': real_cases
    }
    
    # Ensure temp dir exists
    os.makedirs('validation/temp', exist_ok=True)
    with open(TEMP_INPUT_JSON, 'w', encoding='utf-8') as f:
        json.dump(input_payload, f, indent=2)
        
    print("Step 2: Launching TypeScript prediction runner in Vitest...")
    # Run vitest command
    # Note: We need to pass the file path specifically to avoid running other tests.
    res = subprocess.run(
        ["npx", "vitest", "run", "src/run_validation_runner.test.ts"],
        cwd="v2",
        capture_output=True,
        text=True
    )
    
    print(res.stdout)
    if res.returncode != 0:
        print("Error: TypeScript prediction runner failed!")
        print(res.stderr)
        return
        
    print("Step 3: Loading predictions output...")
    if not os.path.exists(TEMP_OUTPUT_JSON):
        print(f"Error: Output predictions file not found at {TEMP_OUTPUT_JSON}!")
        return
        
    with open(TEMP_OUTPUT_JSON, 'r', encoding='utf-8') as f:
        predictions = json.load(f)
        
    ref_preds = predictions.get('reference_predictions', [])
    real_preds = predictions.get('real_predictions', [])
    
    print("Step 4: Writing predictions back to Excel workbook...")
    write_predictions_to_sheet(ref_sheet, ref_preds, ref_mapping)
    write_predictions_to_sheet(real_sheet, real_preds, real_mapping)
    wb.save(EXCEL_PATH)
    print(f"Successfully updated Excel workbook: {EXCEL_PATH}")
    
    # Determine total rows in each sheet (minus header)
    def count_total_rows(ws, case_id_idx):
        count = 0
        for r in range(2, 2000):
            if ws.cell(row=r, column=case_id_idx).value:
                count += 1
            else:
                break
        return count

    ref_total_rows = count_total_rows(ref_sheet, ref_mapping['Case ID'])
    real_total_rows = count_total_rows(real_sheet, real_mapping['Case ID'])
    
    print("Step 5: Calculating metrics and exporting files...")
    # Load library to map organisms to groups
    library_lookup = {}
    if os.path.exists(LIBRARY_JSON_PATH):
        with open(LIBRARY_JSON_PATH, 'r', encoding='utf-8') as f:
            lib_data = json.load(f)
            for bug in lib_data:
                library_lookup[bug['name'].lower()] = bug['group']
                library_lookup[bug['id'].lower()] = bug['group']
    
    ref_metrics = compute_metrics(ref_preds)
    real_metrics = compute_metrics(real_preds)
    
    # Export Results CSVs
    export_results_csv(ref_preds, 'validation/reference_profile_validation_results.csv')
    export_results_csv(real_preds, 'validation/real_case_validation_results.csv')
    
    # Export Metrics CSVs
    export_metrics_csv(ref_metrics, ref_total_rows, 'validation/reference_summary_metrics.csv')
    export_metrics_csv(real_metrics, real_total_rows, 'validation/real_case_summary_metrics.csv')
    
    # Export Confusion Matrices
    generate_organism_confusion_matrix(ref_preds, 'validation/reference_confusion_matrix_organism_level.csv')
    generate_organism_confusion_matrix(real_preds, 'validation/real_case_confusion_matrix_organism_level.csv')
    
    # Group-level confusion matrix for real cases
    if library_lookup:
        generate_group_confusion_matrix(real_preds, library_lookup, 'validation/real_case_confusion_matrix_group_level.csv')
        print("Exported group-level confusion matrix for real cases.")
        
    print("\n" + "="*50)
    print("      BOKBAC DETAILED VALIDATION COMPLETE            ")
    print("="*50)
    print(f"Reference Cases Validated: {ref_metrics['included_cases']}")
    print(f"  Top-1 Accuracy: {round(ref_metrics['top1_accuracy'], 2)}%")
    print(f"  Top-3 Accuracy: {round(ref_metrics['top3_accuracy'], 2)}%")
    print(f"  Macro-F1 score: {round(ref_metrics['macro_f1'], 2)}%")
    print("-"*50)
    print(f"Real Cases Validated:      {real_metrics['included_cases']}")
    print(f"  Top-1 Accuracy: {round(real_metrics['top1_accuracy'], 2)}%")
    print(f"  Top-3 Accuracy: {round(real_metrics['top3_accuracy'], 2)}%")
    print(f"  Macro-F1 score: {round(real_metrics['macro_f1'], 2)}%")
    print("="*50)

if __name__ == '__main__':
    main()
