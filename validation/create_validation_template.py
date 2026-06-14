import os
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

def create_workbook():
    # Paths
    temp_dir = 'validation/temp'
    output_path = 'validation/BokBac_validation_template.xlsx'

    # Load JSON data
    with open(os.path.join(temp_dir, 'tests.json'), 'r', encoding='utf-8') as f:
        tests = json.load(f)
    with open(os.path.join(temp_dir, 'library.json'), 'r', encoding='utf-8') as f:
        library = json.load(f)
    with open(os.path.join(temp_dir, 'suites.json'), 'r', encoding='utf-8') as f:
        suites = json.load(f)

    # Initialize workbook
    wb = openpyxl.Workbook()

    # Style definitions
    font_family = 'Calibri'
    
    title_font = Font(name=font_family, size=16, bold=True, color='1B365D')
    subtitle_font = Font(name=font_family, size=11, italic=True, color='555555')
    section_font = Font(name=font_family, size=12, bold=True, color='1B365D')
    header_font = Font(name=font_family, size=11, bold=True, color='FFFFFF')
    body_font = Font(name=font_family, size=11, color='000000')
    italic_body_font = Font(name=font_family, size=10, italic=True, color='555555')
    bold_body_font = Font(name=font_family, size=11, bold=True, color='000000')
    
    warning_font = Font(name=font_family, size=11, bold=True, color='9C0006')
    warning_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    warning_border = Border(
        left=Side(style='medium', color='9C0006'),
        right=Side(style='medium', color='9C0006'),
        top=Side(style='medium', color='9C0006'),
        bottom=Side(style='medium', color='9C0006')
    )

    header_fill = PatternFill(start_color='1B365D', end_color='1B365D', fill_type='solid')
    meta_header_fill = PatternFill(start_color='2A4D7C', end_color='2A4D7C', fill_type='solid')
    test_header_fill = PatternFill(start_color='3B629B', end_color='3B629B', fill_type='solid')
    output_header_fill = PatternFill(start_color='4C77BA', end_color='4C77BA', fill_type='solid')
    neutral_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

    thin_border_side = Side(style='thin', color='D9D9D9')
    grid_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)

    # Helper: Set gridlines visible
    def set_gridlines(ws):
        ws.views.sheetView[0].showGridLines = True

    # ==========================================
    # SHEET 1: README
    # ==========================================
    ws_readme = wb.active
    ws_readme.title = 'README'
    set_gridlines(ws_readme)

    ws_readme.column_dimensions['A'].width = 3
    ws_readme.column_dimensions['B'].width = 30
    ws_readme.column_dimensions['C'].width = 75

    # Title
    ws_readme['B2'] = 'BokBac Academic Validation Workbook'
    ws_readme['B2'].font = title_font
    ws_readme['B3'] = 'A standardized batch template to validate the Bayesian identification algorithm'
    ws_readme['B3'].font = subtitle_font

    # Warning Box
    ws_readme.merge_cells('B5:C7')
    ws_readme['B5'] = '⚠️ PRIVACY WARNING: DO NOT include patient names, Hospital Numbers (HN), lab numbers, or any other personally identifiable information (PII) in this workbook. All validation cases must be completely de-identified prior to upload.'
    ws_readme['B5'].font = warning_font
    ws_readme['B5'].fill = warning_fill
    ws_readme['B5'].alignment = Alignment(wrap_text=True, vertical='center')
    for r in range(5, 8):
        for c in range(2, 4):
            ws_readme.cell(row=r, column=c).border = warning_border
            if r != 5 or c != 2:
                ws_readme.cell(row=r, column=c).fill = warning_fill

    # Purpose section
    ws_readme['B9'] = '1. Purpose of the Workbook'
    ws_readme['B9'].font = section_font
    ws_readme['B10'] = 'This template is designed to collect validation datasets for the BokBac microbial identification system. By filling this template, users can evaluate the performance of the Bayesian diagnostic engine in batch mode without manual case entry in the UI.'
    ws_readme['B10'].font = body_font
    ws_readme['B10'].alignment = Alignment(wrap_text=True)

    ws_readme['B12'] = '2. Validation Types'
    ws_readme['B12'].font = section_font
    ws_readme['C13'] = '• Reference-Profile Validation: Analytical verification of the scoring logic using ideal organism test profiles (e.g. from textbooks or MCM reference rows) to check if the engine resolves candidate species correctly.'
    ws_readme['C13'].font = body_font
    ws_readme['C13'].alignment = Alignment(wrap_text=True)
    ws_readme['C14'] = '• Real-Case Validation: Evaluation of diagnostic accuracy on de-identified clinical lab cases with known final culture identifications. This is used to compute Top-1 accuracy, Top-3 accuracy, Precision, Recall, F1-scores, and error profiles.'
    ws_readme['C14'].font = body_font
    ws_readme['C14'].alignment = Alignment(wrap_text=True)

    ws_readme['B16'] = '3. How to Fill the Sheets'
    ws_readme['B16'].font = section_font
    ws_readme['C17'] = '• Real_Case_Input: Enter real hospital cases. Fill columns B to M. For tests (columns starting with test__), use the provided dropdown menus to select the observed results. Leave output columns N to T blank—they will be populated by the automated validation runner.'
    ws_readme['C17'].font = body_font
    ws_readme['C17'].alignment = Alignment(wrap_text=True)
    ws_readme['C18'] = '• Reference_Profile_Input: Enter standard reference profiles. Similar to Real_Case_Input, but uses expected_organism instead of final_identification, and includes a reference_source field.'
    ws_readme['C18'].font = body_font
    ws_readme['C18'].alignment = Alignment(wrap_text=True)

    ws_readme['B20'] = '4. Allowed Test Values'
    ws_readme['B20'].font = section_font
    ws_readme['C21'] = '• Binary tests (most tests): + (positive), − (negative), V (variable). Note: Unicode minus sign (U+2212) is the system default.'
    ws_readme['C21'].font = body_font
    ws_readme['C22'] = '• Susceptibility tests (Novobiocin, Optochin, Bacitracin): S (sensitive), R (resistant), V (variable).'
    ws_readme['C22'].font = body_font
    ws_readme['C23'] = '• Hemolysis: β (complete), α (partial/green), γ (none).'
    ws_readme['C23'].font = body_font
    ws_readme['C24'] = '• TSI slant/butt: A/A, A/A (gas+), A/A (gas−), K/A, K/A (gas+), K/A H₂S, K/AG H₂S, K/K, K/N, K/NC.'
    ws_readme['C24'].font = body_font
    ws_readme['C25'] = '• Refer to the Test_Registry sheet to see the exact labels and options for all biochem tests.'
    ws_readme['C25'].font = body_font

    # ==========================================
    # SHEET 6: Allowed_Values (Created first to reference in data validation)
    # ==========================================
    ws_allowed = wb.create_sheet('Allowed_Values')
    set_gridlines(ws_allowed)

    # Header row
    allowed_headers = [
        'Specimen_Types', 'Gram_Reactions', 'Morphologies', 'Arrangements', 
        'Included_Choices', 'Organism_IDs', 'Binary_Options', 'Susceptibility_Options', 
        'Hemolysis_Options', 'TSI_Options', 'CAMP_Options', 'Urease_Options', 
        'Salt6_Options', 'TCBS_Options', 'ColonyXLD_Options', 'Identification_Levels'
    ]

    for col_idx, header in enumerate(allowed_headers, 1):
        cell = ws_allowed.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Values mapping
    allowed_values = {
        'Specimen_Types': ['urine', 'blood', 'stool', 'wound', 'respiratory', 'csf', 'genital', 'throat', 'ear', 'unknown', 'other'],
        'Gram_Reactions': ['positive', 'negative', 'variable', 'unknown'],
        'Morphologies': ['cocci', 'bacilli', 'coccobacilli', 'curved_rod', 'branching_filament', 'unknown'],
        'Arrangements': ['cluster', 'chain', 'pairs', 'diplococci', 'palisade', 'single', 'unknown'],
        'Included_Choices': ['Yes', 'No'],
        'Organism_IDs': sorted([bug['name'] for bug in library]), # Changed to name as requested
        'Binary_Options': ['+', '−', 'V'],
        'Susceptibility_Options': ['S', 'R', 'V'],
        'Hemolysis_Options': ['β (complete)', 'α (partial/green)', 'γ (none)'],
        'TSI_Options': ['A/A', 'A/A (gas+)', 'A/A (gas−)', 'K/A', 'K/A (gas+)', 'K/A H₂S', 'K/AG H₂S', 'K/K', 'K/N', 'K/NC'],
        'CAMP_Options': ['+', '−', 'Reverse CAMP+'],
        'Urease_Options': ['++ (rapid)', '+', '−'],
        'Salt6_Options': ['+', '±', '−'],
        'TCBS_Options': ['Yellow (sucrose+)', 'Green/Blue (sucrose−)', 'No growth'],
        'ColonyXLD_Options': ['Yellow (xylose+)', 'Red (xylose−)', 'Red with black center (H₂S+)'],
        'Identification_Levels': ['species', 'genus', 'complex']
    }

    # Populate lists
    for col_idx, header in enumerate(allowed_headers, 1):
        vals = allowed_values[header]
        for row_idx, val in enumerate(vals, 2):
            ws_allowed.cell(row=row_idx, column=col_idx, value=val).font = body_font

    # Set column widths for Allowed_Values
    for col in ws_allowed.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_allowed.column_dimensions[col_letter].width = max(max_len + 3, 15)

    # Reference cell formulas for validations
    validation_formulas = {
        'specimen_type': f"'Allowed_Values'!$A$2:$A${len(allowed_values['Specimen_Types']) + 1}",
        'gram_reaction': f"'Allowed_Values'!$B$2:$B${len(allowed_values['Gram_Reactions']) + 1}",
        'morphology': f"'Allowed_Values'!$C$2:$C${len(allowed_values['Morphologies']) + 1}",
        'arrangement': f"'Allowed_Values'!$D$2:$D${len(allowed_values['Arrangements']) + 1}",
        'included_in_analysis': f"'Allowed_Values'!$E$2:$E${len(allowed_values['Included_Choices']) + 1}",
        'organism_ids': f"'Allowed_Values'!$F$2:$F${len(allowed_values['Organism_IDs']) + 1}",
        'identification_levels': f"'Allowed_Values'!$P$2:$P${len(allowed_values['Identification_Levels']) + 1}",
        'binary': f"'Allowed_Values'!$G$2:$G${len(allowed_values['Binary_Options']) + 1}",
        'susceptibility': f"'Allowed_Values'!$H$2:$H${len(allowed_values['Susceptibility_Options']) + 1}",
        'hemolysis': f"'Allowed_Values'!$I$2:$I${len(allowed_values['Hemolysis_Options']) + 1}",
        'tsi': f"'Allowed_Values'!$J$2:$J${len(allowed_values['TSI_Options']) + 1}",
        'camp': f"'Allowed_Values'!$K$2:$K${len(allowed_values['CAMP_Options']) + 1}",
        'urease': f"'Allowed_Values'!$L$2:$L${len(allowed_values['Urease_Options']) + 1}",
        'salt_6': f"'Allowed_Values'!$M$2:$M${len(allowed_values['Salt6_Options']) + 1}",
        'salt_8': f"'Allowed_Values'!$M$2:$M${len(allowed_values['Salt6_Options']) + 1}", 
        'tcbs': f"'Allowed_Values'!$N$2:$N${len(allowed_values['TCBS_Options']) + 1}",
        'colony_xld': f"'Allowed_Values'!$O$2:$O${len(allowed_values['ColonyXLD_Options']) + 1}",
    }

    # Hide Allowed_Values sheet
    ws_allowed.sheet_state = 'hidden'

    # ==========================================
    # SHEET 2 & 3: INPUT SHEETS (Real_Case_Input & Reference_Profile_Input)
    # ==========================================
    for is_real in [True, False]:
        sheet_title = 'Real_Case_Input' if is_real else 'Reference_Profile_Input'
        ws = wb.create_sheet(sheet_title)
        set_gridlines(ws)
        ws.freeze_panes = 'A2'

        # Build column definitions
        cols = [
            ('case_id', 'Case ID', meta_header_fill, None),
            ('dataset_type', 'Dataset Type', meta_header_fill, None), # Will configure dv custom values
            ('gram_reaction', 'Gram Reaction', meta_header_fill, validation_formulas['gram_reaction']),
            ('morphology', 'Morphology', meta_header_fill, validation_formulas['morphology']),
            ('arrangement', 'Arrangement', meta_header_fill, validation_formulas['arrangement']),
        ]

        if is_real:
            cols.extend([
                ('specimen_type', 'Specimen Type', meta_header_fill, validation_formulas['specimen_type']),
                ('final_identification', 'Final Identification (ID)', meta_header_fill, validation_formulas['organism_ids']),
                ('identification_level', 'Identification Level', meta_header_fill, validation_formulas['identification_levels']),
                ('identification_method', 'Identification Method', meta_header_fill, None),
            ])
        else:
            cols.extend([
                ('expected_organism', 'Expected Organism (ID)', meta_header_fill, validation_formulas['organism_ids']),
                ('reference_source', 'Reference Source', meta_header_fill, None),
            ])

        cols.extend([
            ('included_in_analysis', 'Included In Analysis', meta_header_fill, validation_formulas['included_in_analysis']),
            ('exclusion_reason', 'Exclusion Reason', meta_header_fill, None),
            ('note', 'Note', meta_header_fill, None),
        ])

        # Biochemical test columns
        for test in tests:
            col_id = f"test__{test['id']}"
            col_label = f"test__{test['id']}"
            
            # Map validation formula
            f_key = test['id']
            if f_key not in validation_formulas:
                f_key = test['resultKind']
            formula = validation_formulas.get(f_key, validation_formulas['binary'])
            
            cols.append((col_id, col_label, test_header_fill, formula))

        # Output columns
        cols.extend([
            ('bokbac_top1', 'bokbac_top1', output_header_fill, None),
            ('bokbac_top2', 'bokbac_top2', output_header_fill, None),
            ('bokbac_top3', 'bokbac_top3', output_header_fill, None),
            ('bokbac_confidence', 'bokbac_confidence', output_header_fill, None),
            ('top1_correct', 'top1_correct', output_header_fill, None),
            ('top3_correct', 'top3_correct', output_header_fill, None),
            ('review_note', 'review_note', output_header_fill, None),
        ])

        # Write Headers
        for col_idx, (col_id, col_label, fill, _) in enumerate(cols, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_label)
            cell.font = header_font
            cell.fill = fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = grid_border

        # Set row height for headers
        ws.row_dimensions[1].height = 28

        # Apply Data Validations & Pre-fill Dataset Type for all 1000 rows
        for col_idx, (col_id, col_label, _, formula) in enumerate(cols, 1):
            col_letter = get_column_letter(col_idx)
            
            if col_id == 'dataset_type':
                # Configure broad valid dataset type choices
                if is_real:
                    allowed_dt = '"real_case,clinical_case,other"'
                    default_dt = 'real_case'
                else:
                    allowed_dt = '"reference_profile,internal_reference_profile,textbook_profile"'
                    default_dt = 'reference_profile'

                dv = DataValidation(type='list', formula1=allowed_dt, allow_blank=False)
                ws.add_data_validation(dv)
                dv.add(f'{col_letter}2:{col_letter}1000')
                
                # Pre-fill dataset_type in all 1000 rows (from row 2 to 1000) for user convenience
                for r in range(2, 1001):
                    ws.cell(row=r, column=col_idx, value=default_dt).font = body_font
            
            elif formula:
                dv = DataValidation(type='list', formula1=formula, allow_blank=True)
                ws.add_data_validation(dv)
                dv.add(f'{col_letter}2:{col_letter}1000')

        # Auto-adjust column widths
        for col_idx, (col_id, col_label, _, _) in enumerate(cols, 1):
            col_letter = get_column_letter(col_idx)
            width = 16
            if col_id in ['case_id', 'dataset_type', 'gram_reaction', 'morphology', 'arrangement']:
                width = 15
            elif col_id in ['exclusion_reason', 'note', 'review_note']:
                width = 25
            ws.column_dimensions[col_letter].width = width

    # ==========================================
    # SHEET 4: Test_Registry
    # ==========================================
    ws_tests = wb.create_sheet('Test_Registry')
    set_gridlines(ws_tests)
    ws_tests.freeze_panes = 'A2'

    test_headers = [
        'test_id', 'test_label', 'category', 'allowed_values', 
        'used_in_diagnostic_workflow', 'notes'
    ]

    for col_idx, header in enumerate(test_headers, 1):
        cell = ws_tests.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = grid_border

    active_test_ids = set()
    for suite in suites:
        for t in suite['tests']:
            active_test_ids.add(t['testId'])

    for row_idx, test in enumerate(tests, 2):
        tid = test['id']
        label = test['label']
        cat = test.get('category', 'general')
        opts_str = ', '.join(test.get('options', []))
        is_active = 'Yes' if tid in active_test_ids else 'No'
        
        note_parts = []
        if test.get('hardExclusion'):
            note_parts.append('Hard Exclusion Gate')
        if test.get('estimatedTime'):
            note_parts.append(f"Time: {test['estimatedTime']}")
        if test.get('costLevel'):
            note_parts.append(f"Cost: {test['costLevel']}")
        note = '; '.join(note_parts)

        ws_tests.cell(row=row_idx, column=1, value=tid).font = bold_body_font
        ws_tests.cell(row=row_idx, column=2, value=label).font = body_font
        ws_tests.cell(row=row_idx, column=3, value=cat).font = body_font
        ws_tests.cell(row=row_idx, column=4, value=opts_str).font = body_font
        ws_tests.cell(row=row_idx, column=5, value=is_active).font = body_font
        ws_tests.cell(row=row_idx, column=6, value=note).font = body_font

        for col in range(1, 7):
            cell = ws_tests.cell(row=row_idx, column=col)
            cell.border = grid_border
            if is_active == 'No':
                cell.fill = neutral_fill

    for col in ws_tests.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_tests.column_dimensions[col_letter].width = max(max_len + 3, 14)

    # ==========================================
    # SHEET 5: Organism_Registry
    # ==========================================
    ws_bugs = wb.create_sheet('Organism_Registry')
    set_gridlines(ws_bugs)
    ws_bugs.freeze_panes = 'C2'

    organism_meta_cols = [
        'organism_id', 'organism_name', 'thai_name', 'genus', 
        'organism_group', 'gram_reaction', 'morphology', 'notes'
    ]
    
    registry_test_ids = [t['id'] for t in tests]
    bug_headers = organism_meta_cols + [f"profile__{tid}" for tid in registry_test_ids]

    for col_idx, header in enumerate(bug_headers, 1):
        cell = ws_bugs.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = meta_header_fill if col_idx <= len(organism_meta_cols) else test_header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = grid_border

    label_to_id = {}
    for t in tests:
        label_to_id[t['label'].lower().replace(' ', '')] = t['id']
        label_to_id[t['id'].lower().replace('_', '')] = t['id']
    
    aliases_override = {
        'h2stsi': 'h2s',
        'motile': 'motility',
        'motility': 'motility',
        'glucoseo/f': 'glucose',
        'pyocyanin': 'king_p',
        'pyoverdin': 'king_f',
        'growth42c': 'growth_42',
        'growth45': 'growth_45',
        'nacl0': 'salt_0',
        'nacl6': 'salt_6',
        'nacl8': 'salt_8',
        'nacl10': 'salt_10',
        'nacl6pct': 'salt_6',
        '6.5%nacl': 'salt_6',
        '6.5nacl': 'salt_6',
        '65nacl': 'salt_6',
        'colonydonxldagar': 'colony_xld',
        'growthmacconkey': 'growth_macconkey',
        'bileesculin': 'bile_esculin',
        'bilesolubility': 'bile_solubility',
    }
    label_to_id.update(aliases_override)

    def find_test_id(name):
        n = name.lower().replace(' ', '').replace('_', '').replace('₂', '2').replace('₃', '3').replace('°', '')
        if n in label_to_id:
            return label_to_id[n]
        return None

    for row_idx, bug in enumerate(library, 2):
        bid = bug['id']
        name = bug['name']
        thai = bug.get('thai', '')
        genus = name.split()[0]
        group = bug.get('group', '')
        gram = bug.get('gram', '')
        morph = bug.get('morph', '')
        note = bug.get('notes', '')

        gram_val = 'positive' if gram == '+' else ('negative' if gram == '-' or gram == '−' else 'variable')

        ws_bugs.cell(row=row_idx, column=1, value=bid).font = bold_body_font
        ws_bugs.cell(row=row_idx, column=2, value=name).font = body_font
        ws_bugs.cell(row=row_idx, column=3, value=thai).font = body_font
        ws_bugs.cell(row=row_idx, column=4, value=genus).font = body_font
        ws_bugs.cell(row=row_idx, column=5, value=group).font = body_font
        ws_bugs.cell(row=row_idx, column=6, value=gram_val).font = body_font
        ws_bugs.cell(row=row_idx, column=7, value=morph).font = body_font
        ws_bugs.cell(row=row_idx, column=8, value=note).font = italic_body_font

        bug_biochem = bug.get('biochem', [])
        profile_map = {}
        for b in bug_biochem:
            tid = find_test_id(b['t'])
            if tid:
                profile_map[tid] = b['r']

        for tid_idx, tid in enumerate(registry_test_ids, len(organism_meta_cols) + 1):
            r_val = profile_map.get(tid, '—')
            cell = ws_bugs.cell(row=row_idx, column=tid_idx, value=r_val)
            cell.font = body_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = grid_border

        for col in range(1, len(bug_headers) + 1):
            ws_bugs.cell(row=row_idx, column=col).border = grid_border

    for col_idx in range(1, len(organism_meta_cols) + 1):
        col_letter = get_column_letter(col_idx)
        width = 16
        if col_idx in [2, 3]:
            width = 24
        elif col_idx == 8:
            width = 30
        ws_bugs.column_dimensions[col_letter].width = width
    for col_idx in range(len(organism_meta_cols) + 1, len(bug_headers) + 1):
        col_letter = get_column_letter(col_idx)
        ws_bugs.column_dimensions[col_letter].width = 14

    wb.save(output_path)
    print(f"Workbook successfully saved to {output_path}")

if __name__ == '__main__':
    create_workbook()
